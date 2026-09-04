# -*- coding: utf-8 -*-
"""
Pruebas del servidor offline (servidor.py).

Levanta el servidor en un puerto libre, envia respuestas por HTTP como lo
haria un celular, y verifica: guardado, deduplicado, estadisticas, panel y
exportacion a Excel con las 49 columnas.

    python test_servidor.py
"""

import json
import os
import tempfile
import threading
import time
import unittest
import urllib.request
import urllib.error

import servidor


def _req(url, method="GET", data=None):
    headers = {}
    body = None
    if data is not None:
        body = json.dumps(data).encode("utf-8")
        headers["Content-Type"] = "application/json"
    r = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        resp = urllib.request.urlopen(r, timeout=5)
        return resp.status, resp.read(), resp.headers
    except urllib.error.HTTPError as e:
        return e.code, e.read(), e.headers


# Dos respuestas de ejemplo (una con premio, una sin premio).
ENTRY_1 = {
    "respondentId": "1001", "nombre": "Ana Ruiz", "empresa": "CEMEX",
    "correo": "ana@example.com", "telefono": "8110000000",
    "nps": 9, "atencion": 10, "organizacion": 8, "lugar": 9, "registro": 7,
    "gustoMas": "Networking / contactos", "gustoMenos": "Horario del evento",
    "beneficios": {"contactos": True, "ideas": True},
    "rol": "Comprador", "perfiles": "Bien adecuados", "numProveedores": 4,
    "comentario": "Excelente evento",
    "emp": [{}, {}, {}],
    "premio": "Cupón de Estacionamiento", "premioLanded": "Cupón de Estacionamiento",
    "dateCreated": "2026-08-31T10:15:00", "dateModified": "2026-08-31T10:15:00",
}
ENTRY_2 = {
    "respondentId": "1002", "nombre": "Beto Lima", "empresa": "VITRO",
    "nps": 6, "rol": "Proveedor",
    "emp": [{"name": "CEMEX", "cal": 8}, {"name": "LAMOSA", "cal": 7}, {}],
    "sugComiteProv": "Más citas", "sugCompradores": "Puntualidad",
    "premio": "", "premioLanded": "Gracias por participar",
    "dateCreated": "2026-08-31T10:20:00", "dateModified": "2026-08-31T10:20:00",
}


class ServidorTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.tmp, "test.db")
        cls.httpd, cls.db = servidor.make_server(0, cls.db_path)  # puerto libre
        cls.port = cls.httpd.server_address[1]
        cls.base = "http://127.0.0.1:%d" % cls.port
        cls.t = threading.Thread(target=cls.httpd.serve_forever, daemon=True)
        cls.t.start()
        time.sleep(0.2)

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()
        cls.httpd.server_close()
        try:
            cls.db.conn.close()
        except Exception:
            pass

    def test_01_health(self):
        code, body, _ = _req(self.base + "/api/health")
        self.assertEqual(code, 200)
        self.assertTrue(json.loads(body)["ok"])

    def test_02_sirve_index(self):
        code, body, hdr = _req(self.base + "/")
        self.assertEqual(code, 200)
        self.assertIn("Expo PyME", body.decode("utf-8"))
        self.assertIn("text/html", hdr.get("Content-Type", ""))

    def test_03_recibir_respuestas(self):
        for e in (ENTRY_1, ENTRY_2):
            code, body, _ = _req(self.base + "/api/respuesta", "POST", e)
            self.assertEqual(code, 200)
            j = json.loads(body)
            self.assertTrue(j["ok"])
            self.assertTrue(j["nuevo"])

    def test_04_dedupe(self):
        # Reenviar la misma respuesta NO debe crear un duplicado.
        code, body, _ = _req(self.base + "/api/respuesta", "POST", ENTRY_1)
        self.assertEqual(code, 200)
        j = json.loads(body)
        self.assertTrue(j["ok"])
        self.assertFalse(j["nuevo"])

    def test_05_stats(self):
        code, body, _ = _req(self.base + "/api/stats")
        self.assertEqual(code, 200)
        s = json.loads(body)
        self.assertEqual(s["total"], 2)     # solo 2 pese al reenvio
        self.assertEqual(s["premios"], 1)   # solo ENTRY_1 tiene premio

    def test_06_json_invalido(self):
        r = urllib.request.Request(
            self.base + "/api/respuesta", data=b"esto no es json",
            headers={"Content-Type": "application/json"}, method="POST")
        try:
            urllib.request.urlopen(r, timeout=5)
            self.fail("debio fallar")
        except urllib.error.HTTPError as e:
            self.assertEqual(e.code, 400)

    def test_07_panel(self):
        code, body, _ = _req(self.base + "/panel")
        self.assertEqual(code, 200)
        self.assertIn("Panel", body.decode("utf-8"))

    def test_08_export_xlsx(self):
        code, body, hdr = _req(self.base + "/export.xlsx")
        self.assertEqual(code, 200)
        self.assertIn("spreadsheetml", hdr.get("Content-Type", ""))
        self.assertEqual(body[:2], b"PK")  # firma de archivo ZIP/XLSX

        # Verifica estructura con openpyxl.
        import openpyxl
        tmp_xlsx = os.path.join(self.tmp, "salida.xlsx")
        with open(tmp_xlsx, "wb") as f:
            f.write(body)
        wb = openpyxl.load_workbook(tmp_xlsx)
        ws = wb.active
        self.assertEqual(ws.max_column, 49)          # 49 columnas exactas
        self.assertEqual(ws.max_row, 4)              # 2 encabezados + 2 datos
        self.assertEqual(ws.cell(1, 1).value, "respondent_id")
        self.assertEqual(ws.cell(1, 9).value, "custom_1")
        # El premio de ENTRY_1 va en custom_1 (columna 9).
        col9 = [ws.cell(r, 9).value for r in range(3, 5)]
        self.assertIn("Cupón de Estacionamiento", col9)

    def test_09_traversal_bloqueado(self):
        # No debe permitir salir de la carpeta img/.
        code, _, _ = _req(self.base + "/img/../servidor.py")
        self.assertIn(code, (403, 404))

    def test_11_config_vacia(self):
        # Sin config guardada, el servidor devuelve null (los clientes usan
        # los valores por defecto del código).
        code, body, _ = _req(self.base + "/api/config")
        self.assertEqual(code, 200)
        self.assertIsNone(json.loads(body)["config"])

    def test_12_config_pin_incorrecto(self):
        cfg = {"prizes": [{"name": "X", "label": "X", "weight": 5,
                           "stock": 10, "isPrize": True}], "pin": "1234"}
        code, _, _ = _req(self.base + "/api/config", "POST",
                          {"pin": "9999", "config": cfg})
        self.assertEqual(code, 403)  # PIN equivocado -> rechazado

    def test_13_config_guardar_y_leer(self):
        cfg = {"prizes": [
                   {"name": "Cupón", "label": "¡PREMIO!", "weight": 20,
                    "stock": 100, "isPrize": True},
                   {"name": "Casi", "label": "¡CASI!", "weight": 80,
                    "stock": None, "isPrize": False}],
               "eventTitle": "Expo PyME", "autoExportEvery": 10, "pin": "1234"}
        code, body, _ = _req(self.base + "/api/config", "POST",
                             {"pin": "1234", "config": cfg})
        self.assertEqual(code, 200)
        self.assertTrue(json.loads(body)["ok"])
        # Se lee de vuelta igual.
        code, body, _ = _req(self.base + "/api/config")
        got = json.loads(body)["config"]
        self.assertEqual(got["prizes"][0]["weight"], 20)
        self.assertEqual(got["pin"], "1234")

    def test_14_config_cambio_de_pin(self):
        # Cambiar el PIN autenticando con el PIN anterior (1234 de test_13).
        cfg = {"prizes": [{"name": "Y", "label": "Y", "weight": 1,
                           "stock": 1, "isPrize": False}], "pin": "4321"}
        code, _, _ = _req(self.base + "/api/config", "POST",
                          {"pin": "1234", "config": cfg})
        self.assertEqual(code, 200)
        # Ahora el PIN válido es 4321; el viejo ya no sirve.
        code, _, _ = _req(self.base + "/api/config", "POST",
                          {"pin": "1234", "config": cfg})
        self.assertEqual(code, 403)

    def test_15_prize_default_2_por_hora(self):
        # Sin config publicada, el cupo por defecto es 2 premios por hora.
        # Los 2 primeros sí se otorgan; el 3º de la misma hora se rechaza.
        for rid in ("t1", "t2"):
            code, body, _ = _req(self.base + "/api/prize", "POST",
                                 {"rid": rid, "premio": "Cupón de Estacionamiento"})
            self.assertEqual(code, 200)
            self.assertTrue(json.loads(body)["claimed"])
        code, body, _ = _req(self.base + "/api/prize", "POST",
                             {"rid": "t3", "premio": "Cupón de Estacionamiento"})
        self.assertEqual(code, 200)
        self.assertFalse(json.loads(body)["claimed"])  # cupo de la hora agotado

    def test_16_prize_status(self):
        code, body, _ = _req(self.base + "/api/prize-status")
        self.assertEqual(code, 200)
        s = json.loads(body)
        self.assertEqual(s["limit"], 2)
        self.assertEqual(s["used"], 2)     # t1 y t2 de la hora actual
        self.assertFalse(s["available"])

    def test_17_prize_sin_rid(self):
        code, _, _ = _req(self.base + "/api/prize", "POST", {"premio": "X"})
        self.assertEqual(code, 400)

    def test_18_prize_ilimitado(self):
        # Configuras 0 = sin límite: el 3º premio de la hora se otorga.
        cfg = {"prizes": [{"name": "Y", "label": "Y", "weight": 1,
                           "stock": 1, "isPrize": False}],
               "prizesPerHour": 0, "pin": "4321"}
        code, body, _ = _req(self.base + "/api/config", "POST",
                             {"pin": "4321", "config": cfg})
        self.assertEqual(code, 200)
        self.assertTrue(json.loads(body)["ok"])
        for rid in ("u1",):
            code, body, _ = _req(self.base + "/api/prize", "POST",
                                 {"rid": rid, "premio": "Cupón"})
            self.assertEqual(code, 200)
            self.assertTrue(json.loads(body)["claimed"])  # sin límite

    def test_10_qr(self):
        code, body, hdr = _req(self.base + "/qr.png")
        if servidor.HAS_QRCODE:
            self.assertEqual(code, 200)
            self.assertIn("image/png", hdr.get("Content-Type", ""))
            self.assertEqual(body[:4], b"\x89PNG")
        else:
            self.assertEqual(code, 404)


class ExportUnitTest(unittest.TestCase):
    """Prueba directa del mapeo a fila de 49 columnas (sin HTTP)."""

    def test_row_49_columnas(self):
        self.assertEqual(len(servidor.HEADER1), 49)
        self.assertEqual(len(servidor.HEADER2), 49)
        self.assertEqual(len(servidor.row_from_entry(ENTRY_1)), 49)
        self.assertEqual(len(servidor.row_from_entry({})), 49)  # entrada vacia

    def test_beneficios_y_emp(self):
        row = servidor.row_from_entry(ENTRY_1)
        # beneficio "contactos" (col index 28) debe traer el texto largo.
        self.assertEqual(row[28], "Acceso a contactos que no obtendría por otra parte")
        # nps en col index 19.
        self.assertEqual(row[19], 9)

    def test_emp_calificaciones(self):
        row = servidor.row_from_entry(ENTRY_2)
        # 1 - Empresa (idx 37), 1 - Calificación (idx 38)
        self.assertEqual(row[37], "CEMEX")
        self.assertEqual(row[38], 8)
        # 2 - Empresa (idx 39), 2 - Calificación (idx 40)
        self.assertEqual(row[39], "LAMOSA")
        self.assertEqual(row[40], 7)


if __name__ == "__main__":
    unittest.main(verbosity=2)
