# -*- coding: utf-8 -*-
"""
Servidor OFFLINE para Expo PyME - Encuesta + Ruleta
===================================================

QUE HACE
--------
Levanta un servidor web LOCAL (sin necesidad de internet) que:
  1) Sirve la pagina index.html a los celulares que se conecten a la red
     local de esta laptop (WiFi/hotspot). Escanean un QR -> abren la encuesta.
  2) Recibe cada respuesta por HTTP POST y la guarda de forma CENTRALIZADA en
     una base de datos SQLite (respuestas.db) en ESTA laptop.
  3) Ofrece un panel (/panel) con el QR, el conteo en vivo y el boton para
     exportar el Excel con el MISMO formato de 49 columnas de siempre.

NO necesita internet en ningun momento. Solo requiere que los celulares esten
conectados a la MISMA red WiFi que esta laptop (por ejemplo el "Mobile Hotspot"
de Windows, cuya IP de host siempre es 192.168.137.1).

USO
---
    python servidor.py                # puerto 8080 por defecto
    python servidor.py --port 8080    # elegir puerto
    python servidor.py --db mis.db    # elegir archivo de base de datos

Luego, en la terminal veras la URL y un QR para escanear. Abre el panel en
esta laptop:  http://localhost:8080/panel

REQUISITOS
----------
- Python 3.8+ (incluye http.server y sqlite3, no requieren instalacion).
- openpyxl  -> para exportar a Excel (pip install openpyxl).
- qrcode    -> OPCIONAL, solo para dibujar el QR (pip install qrcode).
              Si no esta, el panel muestra la URL en grande igualmente.

Instala lo necesario UNA sola vez, con internet, antes del evento:
    pip install openpyxl qrcode
"""

import argparse
import io
import json
import mimetypes
import os
import re
import socket
import sqlite3
import sys
import threading
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import unquote

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INDEX_FILE = os.path.join(BASE_DIR, "index.html")

# Extensiones estaticas permitidas (seguridad: no servir cualquier archivo)
STATIC_DIRS = ("img",)

# Rutas con que Android/iOS/Windows detectan "¿hay internet?" al conectarse a
# una red. Si respondemos con una redireccion en vez de lo que esperan, el
# sistema abre solo la pagina (captive portal). OJO: esto SOLO se dispara si
# esta laptop es el punto de acceso o el DNS de la red; si la laptop es solo un
# cliente de un hotspot de celular, el celular nunca manda estas peticiones
# aqui. Ver README / nota de despliegue.
CAPTIVE_PROBE_PATHS = (
    "/generate_204", "/gen_204",           # Android
    "/hotspot-detect.html",                # iOS / macOS
    "/library/test/success.html",          # iOS (variante)
    "/connecttest.txt", "/ncsi.txt",       # Windows (NCSI)
    "/redirect",                           # Windows
    "/canonical.html",                     # Firefox / GNOME
    "/success.txt",                        # Firefox
    "/kindle-wifi/wifistub.html",          # Kindle
)


def wifi_qr_payload(ssid, password, enc="WPA"):
    """Texto estandar para un QR que conecta el telefono a una red WiFi.
    Formato: WIFI:S:<ssid>;T:<WPA|nopass>;P:<clave>;;  Escapa \\ ; , : \" ."""
    def esc(s):
        return re.sub(r'([\\;,:"])', r"\\\1", s or "")
    enc = (enc or "WPA").upper()
    if enc in ("NOPASS", "NONE", ""):
        return "WIFI:S:%s;T:nopass;;" % esc(ssid)
    return "WIFI:S:%s;T:%s;P:%s;;" % (esc(ssid), enc, esc(password))

# ---- Dependencias opcionales -------------------------------------------------
try:
    import openpyxl  # noqa: F401
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

try:
    import qrcode
    HAS_QRCODE = True
except Exception:
    HAS_QRCODE = False


# =============================================================================
# BASE DE DATOS (SQLite)
# =============================================================================
class DB:
    """Acceso concurrente seguro a SQLite (varios celulares a la vez)."""

    def __init__(self, path):
        self.path = path
        self.lock = threading.Lock()
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self.conn.execute("PRAGMA journal_mode=WAL;")
        self.conn.execute(
            """CREATE TABLE IF NOT EXISTS respuestas(
                   id            INTEGER PRIMARY KEY AUTOINCREMENT,
                   respondent_id TEXT UNIQUE,
                   payload       TEXT NOT NULL,
                   premio        TEXT,
                   created_at    TEXT,
                   recibido_en   TEXT,
                   ip            TEXT
               )"""
        )
        # Configuración central (ruleta/probabilidades). Una sola fila.
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS app_config(k TEXT PRIMARY KEY, v TEXT)"
        )
        self.conn.commit()

    def get_config(self):
        with self.lock:
            row = self.conn.execute(
                "SELECT v FROM app_config WHERE k='config'").fetchone()
        if not row:
            return None
        try:
            return json.loads(row[0])
        except Exception:
            return None

    def set_config(self, obj):
        payload = json.dumps(obj, ensure_ascii=False)
        with self.lock:
            self.conn.execute(
                "INSERT INTO app_config(k,v) VALUES('config',?) "
                "ON CONFLICT(k) DO UPDATE SET v=excluded.v", (payload,))
            self.conn.commit()

    def insert(self, entry, ip):
        """Inserta una respuesta. Idempotente por respondent_id (evita duplicados
        si un celular reintenta el envio)."""
        rid = str(entry.get("respondentId") or "")
        payload = json.dumps(entry, ensure_ascii=False)
        premio = entry.get("premio") or ""
        created = entry.get("dateCreated") or ""
        recibido = datetime.now().isoformat(timespec="seconds")
        with self.lock:
            cur = self.conn.execute(
                """INSERT OR IGNORE INTO respuestas
                   (respondent_id,payload,premio,created_at,recibido_en,ip)
                   VALUES (?,?,?,?,?,?)""",
                (rid, payload, premio, created, recibido, ip),
            )
            self.conn.commit()
            nuevo = cur.rowcount > 0
        return nuevo

    def stats(self):
        with self.lock:
            total = self.conn.execute("SELECT COUNT(*) FROM respuestas").fetchone()[0]
            premios = self.conn.execute(
                "SELECT COUNT(*) FROM respuestas WHERE premio<>''"
            ).fetchone()[0]
        return {"total": total, "premios": premios}

    def entries(self):
        with self.lock:
            rows = self.conn.execute(
                "SELECT payload FROM respuestas ORDER BY id"
            ).fetchall()
        out = []
        for (payload,) in rows:
            try:
                out.append(json.loads(payload))
            except Exception:
                pass
        return out


# =============================================================================
# EXPORTACION A EXCEL  (mismo layout de 49 columnas que SurveyMonkey)
# =============================================================================
HEADER1 = [
    "respondent_id", "collector_id", "date_created", "date_modified", "ip_address",
    "email_address", "first_name", "last_name", "custom_1",
    "Información de contacto (confidencial)", "", "", "", "", "", "", "", "", "",
    "En una escala del 0 al 10  ¿Qué probabilidad hay de que recomiende este evento?",
    "Favor de evaluar los siguientes aspectos del evento donde 1 es \"muy malo\" y 10 \"excelente\"",
    "", "", "",
    "¿Qué fue lo que más le gustó de este evento?", "",
    "¿Qué fue lo que menos le gustó de este evento?", "",
    "¿Cuáles considera que son los beneficios que adquirió al participar en este evento?",
    "", "", "", "", "", "",
    "¿Cuál fue su rol de participación en el evento?",
    "¿Desea agregar algún comentario o sugerencia?",
    "Indica con qué empresas compradoras tuviste una mejor experiencia, y califica esa "
    "experiencia en una escala del 1 al 10 (donde 1 es \"muy mala\" y 10 es \"excelente\").",
    "", "", "", "", "",
    "¿Qué sugerencias de mejora le darías a cada uno de los siguientes?", "",
    "¿En que medida considera que los perfiles con los que interactuó en el evento, "
    "fueron los adecuados para sus requerimientos?",
    "¿Cuántos de los proveedores que atendiste hoy, te parecieron que podrían ser "
    "potenciales? (Número aproximado)",
    "¿Qué sugerencias de mejora le darías a cada uno de los siguientes?", "",
]
HEADER2 = [
    "", "", "", "", "", "", "", "", "",
    "Nombre", "Empresa", "Dirección", "Dirección 2", "Ciudad/Localidad",
    "Estado/Provincia", "Código postal", "País", "Correo electrónico", "Número de teléfono",
    "Response",
    "Atención del personal", "Organización", "Lugar del evento", "Registro del evento",
    "Response", "Otro (especifique)",
    "Response", "Otro (especifique)",
    "Acceso a contactos que no obtendría por otra parte", "Experiencia",
    "Habilidades empresariales", "Información sobre el mercado y competidores",
    "Ideas de innovación", "Crecimiento para mi empresa", "Otro (especifique)",
    "Response",
    "Open-Ended Response",
    "1 - Empresa", "1 - Calificación", "2 - Empresa", "2 - Calificación",
    "3 - Empresa", "3 - Calificación",
    "Al comité organizador del evento", "A los compradores que te atendieron hoy",
    "Response",
    "Open-Ended Response",
    "Al comité organizador del evento", "A los proveedores que atendiste hoy",
]


def _num(v):
    if v == "" or v is None:
        return ""
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return v


def _ben(e, k, t):
    b = e.get("beneficios") or {}
    return t if b.get(k) else ""


def _emp(e, i, f):
    arr = e.get("emp") or []
    x = arr[i] if i < len(arr) and isinstance(arr[i], dict) else {}
    v = x.get(f)
    return v if v is not None else ""


def _date(iso):
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(str(iso).replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return iso


def row_from_entry(e):
    """Convierte una respuesta (dict) en una fila de 49 celdas. Espeja
    exactamente rowFromEntry() del index.html."""
    return [
        e.get("respondentId") or "", "", _date(e.get("dateCreated")),
        _date(e.get("dateModified")), "", "", "", "", e.get("premio") or "",
        e.get("nombre") or "", e.get("empresa") or "", e.get("direccion") or "",
        e.get("direccion2") or "", e.get("ciudad") or "", e.get("estado") or "",
        e.get("cp") or "", e.get("pais") or "", e.get("correo") or "", e.get("telefono") or "",
        _num(e.get("nps")),
        _num(e.get("atencion")), _num(e.get("organizacion")), _num(e.get("lugar")),
        _num(e.get("registro")),
        e.get("gustoMas") or "", e.get("gustoMasOtro") or "",
        e.get("gustoMenos") or "", e.get("gustoMenosOtro") or "",
        _ben(e, "contactos", "Acceso a contactos que no obtendría por otra parte"),
        _ben(e, "experiencia", "Experiencia"),
        _ben(e, "habilidades", "Habilidades empresariales"),
        _ben(e, "informacion", "Información sobre el mercado y competidores"),
        _ben(e, "ideas", "Ideas de innovación"),
        _ben(e, "crecimiento", "Crecimiento para mi empresa"),
        e.get("beneficioOtro") or "",
        e.get("rol") or "",
        e.get("comentario") or "",
        _emp(e, 0, "name"), _num(_emp(e, 0, "cal")),
        _emp(e, 1, "name"), _num(_emp(e, 1, "cal")),
        _emp(e, 2, "name"), _num(_emp(e, 2, "cal")),
        e.get("sugComiteProv") or "", e.get("sugCompradores") or "",
        e.get("perfiles") or "",
        e.get("numProveedores") if e.get("numProveedores") is not None else "",
        e.get("sugComiteComp") or "", e.get("sugProveedores") or "",
    ]


def build_xlsx(entries):
    """Devuelve los bytes de un .xlsx (49 columnas, 2 filas de encabezado)."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no esta instalado (pip install openpyxl)")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(HEADER1)
    ws.append(HEADER2)
    date_fmt = "mm/dd/yyyy hh:mm:ss"
    for e in entries:
        row = row_from_entry(e)
        ws.append(row)
        r = ws.max_row
        for ci in (3, 4):  # date_created, date_modified
            cell = ws.cell(row=r, column=ci)
            if isinstance(cell.value, datetime):
                cell.number_format = date_fmt
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# RED: detectar IP de la red local
# =============================================================================
def _usable(ip):
    """Descarta direcciones que ningun celular puede alcanzar."""
    if not ip:
        return False
    if ip.startswith("127."):          # loopback local
        return False
    if ip.startswith("169.254."):      # APIPA (sin DHCP / desconectada)
        return False
    if ip.startswith("192.168.199."):  # nuestro adaptador loopback (inutil como AP)
        return False
    return True


def lan_ips():
    """Devuelve las IPv4 candidatas de esta laptop. La PRIMERA es la de la red
    activa (la que tiene salida por la ruta por defecto = la del hotspot/router
    al que esta conectada la laptop). El resto van como alternativas."""
    result = []

    # 1) IP de la interfaz activa (la que usa la ruta por defecto). Funciona aun
    #    sin internet: basta con que el hotspot/router de un gateway por DHCP.
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        primary = s.getsockname()[0]
        s.close()
        if _usable(primary):
            result.append(primary)
    except Exception:
        pass

    # 2) Otras direcciones del equipo, como alternativas.
    others = set()
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
            others.add(info[4][0])
    except Exception:
        pass

    def rank(ip):
        if ip.startswith("192.168.137."):  # Mobile Hotspot de Windows
            return 0
        if ip.startswith("172.20.10."):    # hotspot de iPhone
            return 1
        if ip.startswith("192.168."):
            return 2
        if ip.startswith("10.") or ip.startswith("172."):
            return 3
        return 4

    for ip in sorted(others, key=rank):
        if _usable(ip) and ip not in result:
            result.append(ip)

    return result or ["127.0.0.1"]


def primary_url(port):
    return "http://%s:%d" % (lan_ips()[0], port)


def qr_png_bytes(url):
    if not HAS_QRCODE:
        return None
    qr = qrcode.QRCode(border=2, box_size=10)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def qr_ascii(url):
    if not HAS_QRCODE:
        return ""
    qr = qrcode.QRCode(border=1)
    qr.add_data(url)
    qr.make(fit=True)
    buf = io.StringIO()
    qr.print_ascii(out=buf, invert=True)
    return buf.getvalue()


# =============================================================================
# PANEL (dashboard servido en la laptop)
# =============================================================================
def panel_html(port, wifi=None):
    url = primary_url(port)
    img_style = ('width:260px;height:260px;image-rendering:pixelated;'
                 'border:10px solid #fff;border-radius:12px;background:#fff')
    if not HAS_QRCODE:
        qr_tag = ('<p style="color:#e11b22">Instala <code>qrcode</code> para ver el QR aquí '
                  '(pip install qrcode). La URL de abajo funciona igual.</p>')
    elif wifi:
        # Dos QR: (1) unirse al WiFi, (2) abrir la encuesta.
        qr_tag = (
            '<div style="display:flex;flex-wrap:wrap;gap:24px;justify-content:center">'
            '  <div style="max-width:280px">'
            '    <div style="font-weight:600;margin-bottom:8px">1. Conéctate al WiFi</div>'
            '    <img src="/wifi-qr.png" alt="QR WiFi" style="%s">'
            '    <div class="muted" style="margin-top:6px">Red: <b>%s</b></div>'
            '  </div>'
            '  <div style="max-width:280px">'
            '    <div style="font-weight:600;margin-bottom:8px">2. Abre la encuesta</div>'
            '    <img src="/qr.png" alt="QR encuesta" style="%s">'
            '  </div>'
            '</div>' % (img_style, wifi.get("ssid", ""), img_style))
    else:
        qr_tag = '<img src="/qr.png" alt="QR" style="%s">' % img_style
    return """<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Panel · Expo PyME</title>
<style>
  :root{--brand:#8c1d40;--brand-2:#e11b22}
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif;
       background:#f3eff0;color:#1b1e23;line-height:1.5}
  .wrap{max-width:920px;margin:0 auto;padding:28px 18px 80px}
  h1{font-weight:300;letter-spacing:-.02em;margin:0 0 4px}
  .muted{color:#868d95;margin:0 0 24px}
  .grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
  .card{background:#fff;border:1px solid #e6e9ed;border-radius:18px;padding:22px;
        box-shadow:0 10px 30px rgba(27,30,35,.05)}
  .stat{text-align:center}
  .stat b{display:block;font-size:52px;font-weight:300;line-height:1}
  .stat span{font-size:12px;text-transform:uppercase;letter-spacing:.12em;color:#9aa1a9}
  .qr{display:flex;flex-direction:column;align-items:center;text-align:center;gap:10px}
  .url{font-size:22px;font-weight:600;color:var(--brand);word-break:break-all}
  .btn{display:inline-block;border:none;border-radius:999px;padding:14px 26px;font-weight:600;
       font-size:15px;cursor:pointer;text-decoration:none}
  .btn-dark{background:var(--brand);color:#fff}
  .row{display:flex;gap:12px;flex-wrap:wrap;align-items:center}
  code{background:#f0ebed;padding:2px 6px;border-radius:6px}
  @media(max-width:640px){.grid{grid-template-columns:1fr}.stat b{font-size:40px}}
</style></head><body>
<div class="wrap">
  <h1>Panel de control · Expo PyME</h1>
  <p class="muted">Servidor local activo. Actualiza cada 4&nbsp;s.</p>
  <div class="grid">
    <div class="card stat"><b id="total">0</b><span>Respuestas recibidas</span></div>
    <div class="card stat"><b id="premios">0</b><span>Premios entregados</span></div>
  </div>
  <div class="card qr">
    <h2 style="margin:0">Escanea para responder</h2>
    __QR__
    <div class="url">__URL__</div>
    <p class="muted" style="margin:4px 0 0">Los celulares deben estar en la misma red WiFi de esta laptop.</p>
  </div>
  <div class="card" style="margin-top:20px">
    <div class="row">
      <a class="btn btn-dark" href="/export.xlsx">⬇ Exportar Excel (49 columnas)</a>
      <a class="btn btn-dark" href="/" target="_blank" style="background:#5a6067">Abrir encuesta</a>
    </div>
    <p class="muted" style="margin:14px 0 0">Base de datos: <code>respuestas.db</code> en la carpeta del servidor.
       Exporta con frecuencia como respaldo.</p>
  </div>
</div>
<script>
async function tick(){
  try{
    const r=await fetch('/api/stats',{cache:'no-store'});
    const s=await r.json();
    document.getElementById('total').textContent=s.total;
    document.getElementById('premios').textContent=s.premios;
  }catch(e){}
}
tick(); setInterval(tick,4000);
</script>
</body></html>""".replace("__QR__", qr_tag).replace("__URL__", url)


# =============================================================================
# SERVIDOR HTTP
# =============================================================================
def make_handler(db, port, wifi=None):
    # Hosts "propios": si el Host de la peticion no es uno de estos, asumimos
    # que es una sonda de captive portal (el SO preguntando por otro dominio) y
    # redirigimos a la encuesta. Se calcula una vez al arrancar.
    known_hosts = {"localhost", "127.0.0.1", "0.0.0.0"}
    known_hosts.update(lan_ips())

    class Handler(BaseHTTPRequestHandler):
        server_version = "ExpoPyME/1.0"
        protocol_version = "HTTP/1.1"

        # ---- utilidades de respuesta ----
        def _send(self, code, body, ctype="text/html; charset=utf-8", extra=None):
            if isinstance(body, str):
                body = body.encode("utf-8")
            self.send_response(code)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            for k, v in (extra or {}).items():
                self.send_header(k, v)
            self.end_headers()
            if self.command != "HEAD":
                self.wfile.write(body)

        def _json(self, obj, code=200):
            self._send(code, json.dumps(obj, ensure_ascii=False),
                       "application/json; charset=utf-8")

        def _redirect(self, url, code=302):
            self.send_response(code)
            self.send_header("Location", url)
            self.send_header("Content-Length", "0")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()

        def _is_captive_probe(self, path):
            """True si el SO esta preguntando '¿hay internet?' (para abrirle la
            encuesta como captive portal). Reconoce las rutas conocidas y
            tambien cualquier peticion a un Host que no es este servidor."""
            if path in CAPTIVE_PROBE_PATHS:
                return True
            host = (self.headers.get("Host") or "").split(":")[0].strip().lower()
            return bool(host) and host not in known_hosts

        def log_message(self, fmt, *args):
            pass  # silencio (evita saturar la terminal en el evento)

        # ---- GET ----
        def do_GET(self):
            path = unquote(self.path.split("?", 1)[0])
            # Captive portal: si el SO sondea por internet, lo mandamos a la
            # encuesta. (Solo aplica si esta laptop es el AP/DNS de la red.)
            if path != "/" and self._is_captive_probe(path):
                return self._redirect(primary_url(port) + "/")
            if path in ("/", "/index.html"):
                return self._serve_file(INDEX_FILE, "text/html; charset=utf-8")
            if path == "/panel":
                return self._send(200, panel_html(port, wifi))
            if path == "/qr.png":
                png = qr_png_bytes(primary_url(port))
                if png is None:
                    return self._send(404, "qrcode no instalado")
                return self._send(200, png, "image/png")
            if path == "/wifi-qr.png":
                if not wifi:
                    return self._send(404, "wifi no configurado")
                png = qr_png_bytes(wifi_qr_payload(
                    wifi.get("ssid", ""), wifi.get("pass", ""), wifi.get("enc", "WPA")))
                if png is None:
                    return self._send(404, "qrcode no instalado")
                return self._send(200, png, "image/png")
            if path == "/api/health":
                return self._json({"ok": True})
            if path == "/api/stats":
                return self._json(db.stats())
            if path == "/api/config":
                return self._json({"config": db.get_config()})
            if path == "/api/entries":
                return self._json(db.entries())
            if path == "/export.xlsx":
                return self._export()
            # estaticos (img/...)
            if self._is_static(path):
                return self._serve_static(path)
            return self._send(404, "No encontrado")

        do_HEAD = do_GET

        # ---- POST ----
        def do_POST(self):
            path = self.path.split("?", 1)[0]
            if path == "/api/respuesta":
                return self._recibir()
            if path == "/api/config":
                return self._set_config()
            return self._json({"ok": False, "error": "ruta desconocida"}, 404)

        # ---- handlers concretos ----
        def _recibir(self):
            try:
                length = int(self.headers.get("Content-Length") or 0)
                raw = self.rfile.read(length) if length else b"{}"
                entry = json.loads(raw.decode("utf-8"))
            except Exception as ex:
                return self._json({"ok": False, "error": "json invalido: %s" % ex}, 400)
            if not isinstance(entry, dict):
                return self._json({"ok": False, "error": "se esperaba un objeto"}, 400)
            ip = self.client_address[0] if self.client_address else ""
            try:
                nuevo = db.insert(entry, ip)
            except Exception as ex:
                return self._json({"ok": False, "error": str(ex)}, 500)
            return self._json({"ok": True, "nuevo": nuevo, "total": db.stats()["total"]})

        def _set_config(self):
            """Guarda la configuración central (ruleta/probabilidades). Protegido
            por PIN: el body debe traer el PIN vigente."""
            try:
                length = int(self.headers.get("Content-Length") or 0)
                raw = self.rfile.read(length) if length else b"{}"
                body = json.loads(raw.decode("utf-8"))
            except Exception as ex:
                return self._json({"ok": False, "error": "json invalido: %s" % ex}, 400)
            cfg = body.get("config")
            pin = str(body.get("pin") or "")
            if not isinstance(cfg, dict):
                return self._json({"ok": False, "error": "config invalida"}, 400)
            actual = db.get_config()
            esperado = str((actual or {}).get("pin") or "1234")
            if pin != esperado:
                return self._json({"ok": False, "error": "pin incorrecto"}, 403)
            db.set_config(cfg)
            return self._json({"ok": True})

        def _export(self):
            if not HAS_OPENPYXL:
                return self._send(500, "openpyxl no instalado (pip install openpyxl)")
            try:
                data = build_xlsx(db.entries())
            except Exception as ex:
                return self._send(500, "Error al exportar: %s" % ex)
            stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
            fn = "Encuesta de satisfaccion Expo PyME %s.xlsx" % stamp
            self._send(
                200, data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                {"Content-Disposition": 'attachment; filename="%s"' % fn},
            )

        # ---- archivos ----
        def _is_static(self, path):
            p = path.lstrip("/")
            return any(p.startswith(d + "/") for d in STATIC_DIRS)

        def _serve_static(self, path):
            rel = os.path.normpath(path.lstrip("/")).replace("\\", "/")
            full = os.path.join(BASE_DIR, rel)
            # Evita salir de la carpeta base o de las carpetas estáticas
            # permitidas (path traversal, p.ej. /img/../servidor.py).
            if not os.path.abspath(full).startswith(BASE_DIR):
                return self._send(403, "Prohibido")
            if not any(rel == d or rel.startswith(d + "/") for d in STATIC_DIRS):
                return self._send(403, "Prohibido")
            ctype = mimetypes.guess_type(full)[0] or "application/octet-stream"
            return self._serve_file(full, ctype)

        def _serve_file(self, full, ctype):
            try:
                with open(full, "rb") as f:
                    data = f.read()
            except FileNotFoundError:
                return self._send(404, "Archivo no encontrado")
            return self._send(200, data, ctype)

    return Handler


def make_server(port, db_path, wifi=None):
    db = DB(db_path)
    httpd = ThreadingHTTPServer(("0.0.0.0", port), make_handler(db, port, wifi))
    httpd.daemon_threads = True
    return httpd, db


def main():
    ap = argparse.ArgumentParser(description="Servidor offline Expo PyME (encuesta + ruleta).")
    ap.add_argument("--port", type=int, default=8080, help="Puerto (default 8080)")
    ap.add_argument("--db", default=os.path.join(BASE_DIR, "respuestas.db"),
                    help="Archivo de base de datos SQLite")
    ap.add_argument("--ssid", default="",
                    help="Nombre (SSID) de la red WiFi. Si lo das, el panel muestra "
                         "un 2do QR para que los celulares se unan a la red al escanearlo.")
    ap.add_argument("--wifi-pass", dest="wifi_pass", default="",
                    help="Clave de la red WiFi (para el QR de WiFi).")
    ap.add_argument("--wifi-enc", dest="wifi_enc", default="WPA",
                    help="Cifrado del WiFi: WPA (default) o nopass si es red abierta.")
    args = ap.parse_args()

    wifi = None
    if args.ssid:
        wifi = {"ssid": args.ssid, "pass": args.wifi_pass, "enc": args.wifi_enc}

    # La consola de Windows suele ser cp1252 y no puede dibujar el QR de bloques
    # ni acentos. Forzamos UTF-8 en la salida (no afecta la red ni los datos).
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass

    httpd, _ = make_server(args.port, args.db, wifi)
    url = primary_url(args.port)

    line = "=" * 60
    print(line)
    print("  EXPO PYME - Servidor OFFLINE activo")
    print(line)
    print("  Base de datos : %s" % args.db)
    print("  openpyxl (Excel): %s" % ("OK" if HAS_OPENPYXL else "FALTA -> pip install openpyxl"))
    print("  qrcode (QR)     : %s" % ("OK" if HAS_QRCODE else "FALTA -> pip install qrcode"))
    print(line)
    print("  1) En esta laptop enciende el 'Mobile Hotspot' de Windows.")
    print("  2) Los celulares se conectan a esa red WiFi (no necesitan datos).")
    print("  3) Escanean este QR o escriben la URL en el navegador:")
    print()
    if HAS_QRCODE:
        try:
            print(qr_ascii(url))
        except Exception:
            print("  (No se pudo dibujar el QR aquí; ábrelo en el panel: %s/panel)" % url)
    print("     >>>  %s  <<<" % url)
    print()
    ips = lan_ips()
    if len(ips) > 1:
        print("  Si el QR no abre, prueba otra IP de esta laptop:")
        for ip in ips:
            print("       http://%s:%d" % (ip, args.port))
        print()
    if wifi:
        print("  QR de WiFi ACTIVO en el panel (red '%s'):" % wifi["ssid"])
        print("     1) escanean el QR de WiFi  -> se unen a la red")
        print("     2) escanean el QR de abajo -> abre la encuesta")
        print()
    print("  PANEL de control (en esta laptop): %s/panel" % url)
    print("  Exportar Excel:                    %s/export.xlsx" % url)
    print(line)
    print("  Presiona Ctrl+C para detener el servidor.")
    print(line)
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nDeteniendo servidor...")
        httpd.shutdown()


if __name__ == "__main__":
    main()
