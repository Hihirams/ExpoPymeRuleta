# -*- coding: utf-8 -*-
"""
Descargador de encuestas SurveyMonkey -> Excel (formato export manual) - CAINTRA
================================================================================

QUE HACE
--------
Baja por API las respuestas de UN recopilador (un evento, ej. DC240626) y
genera un archivo .xlsx con EXACTAMENTE el mismo layout de columnas que el
export manual de SurveyMonkey. Asi tu script 'sistema_encuestas_CAINTRA.py'
lo procesa sin cambios: reemplaza el paso de "exportar a mano y copiar el
archivo" por un comando.

NO modifica nada en SurveyMonkey (solo lee) y NO toca tu script de produccion.

REQUISITOS
----------
- Archivo 'sm_token.txt' en la misma carpeta con tu Access Token (una linea).
- pip install requests openpyxl

USO
---
1) Ver los recopiladores (eventos) de una encuesta:
     python descargar_encuestas.py 411209771 --collectors

2) Generar el Excel de un evento (queda en esta carpeta con el nombre que
   tu script espera, ej. "Encuesta de satisfaccion Eventos.xlsx"):
     python descargar_encuestas.py 411209771 DC240626

   Para elegir el nombre/ruta de salida:
     python descargar_encuestas.py 411209771 DC240626 --salida "C:\\ruta\\archivo.xlsx"

3) VERIFICAR contra un export manual existente (recomendado la 1a vez):
     python descargar_encuestas.py 411209771 DC240626 --verificar "Encuesta de satisfaccion Eventos.xlsx"
   Genera el Excel por API y lo compara celda por celda contra tu export
   manual. Te dice si el layout y los datos coinciden 100%.

4) (Avanzado) Construir desde un dump local sin llamar al API:
     python descargar_encuestas.py 411209771 DC240626 --desde-dump
   Usa dump_details_<survey>.json y dump_responses_<collector>.json.
"""

import sys
import os
import re
import json

try:
    import requests
except ImportError:
    print("Falta 'requests'. Instala con:  pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Falta 'openpyxl'. Instala con:  pip install openpyxl")
    sys.exit(1)

from datetime import datetime, timedelta

API_BASE = "https://api.surveymonkey.com/v3"
CARPETA = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_TOKEN = os.path.join(CARPETA, "sm_token.txt")

# ----------------------------------------------------------------------------
# TIPOS DE ENCUESTA (para el menu). El nombre de archivo de salida es EL MISMO
# que espera tu script sistema_encuestas_CAINTRA.py, para que puedas correr el
# script sin cambiar nada. Si algun ID cambia, edítalo aqui.
# ----------------------------------------------------------------------------
TIPOS_ENCUESTA = [
    {"nombre": "Satisfacción de Eventos (presencial)", "survey_id": "411209771",
     "salida": "Encuesta de satisfacción Eventos.xlsx"},
    {"nombre": "Satisfacción de Módulos", "survey_id": "526014848",
     "salida": "Encuesta de satisfacción módulos.xlsx"},
    {"nombre": "Satisfacción para Patrocinadores", "survey_id": "417275970",
     "salida": "Encuesta de satisfacción para patrocinadores.xlsx"},
    {"nombre": "Encuentros de negocios", "survey_id": "421062044",
     "salida": "Encuesta de satisfacción de Encuentros de negocios.xlsx"},
    {"nombre": "Satisfacción Virtual (Zoom)", "survey_id": "",
     "salida": "zoom.xlsx"},
]

# El API entrega fechas en UTC; el export manual de SurveyMonkey las muestra en
# la zona horaria de la cuenta (Monterrey = UTC-6, sin horario de verano).
# Si algun dia cambia la zona de la cuenta, ajusta este valor.
OFFSET_HORAS = -6


# ----------------------------------------------------------------------------
# Token / sesion
# ----------------------------------------------------------------------------
def cargar_token():
    tok = os.environ.get("SM_ACCESS_TOKEN")
    if tok:
        return tok.strip()
    if not os.path.exists(ARCHIVO_TOKEN):
        print("No encontre 'sm_token.txt' en %s" % CARPETA)
        sys.exit(1)
    with open(ARCHIVO_TOKEN, "r", encoding="utf-8") as f:
        tok = f.read().strip()
    if not tok:
        print("'sm_token.txt' esta vacio.")
        sys.exit(1)
    return tok


def crear_sesion(token):
    s = requests.Session()
    s.headers.update({
        "Authorization": "bearer %s" % token,
        "Content-Type": "application/json",
    })
    return s


def _get(s, url):
    r = s.get(url, timeout=90)
    if r.status_code != 200:
        print("ERROR %s en %s\n%s" % (r.status_code, url, r.text))
        sys.exit(1)
    return r.json()


# ----------------------------------------------------------------------------
# Recopiladores
# ----------------------------------------------------------------------------
def obtener_recopiladores(s, survey_id):
    data = _get(s, "%s/surveys/%s/collectors?per_page=1000" % (API_BASE, survey_id))
    return data.get("data", [])


def listar_recopiladores(s, survey_id):
    cols = obtener_recopiladores(s, survey_id)
    print("\nRecopiladores de la encuesta %s (%d):\n" % (survey_id, len(cols)))
    for c in cols:
        print("   %-14s  %s" % (c.get("id"), c.get("name")))


def resolver_collector_id(s, survey_id, nombre_o_id):
    cols = obtener_recopiladores(s, survey_id)
    for c in cols:
        if str(c.get("id")) == str(nombre_o_id):
            return c.get("id"), c.get("name")
    objetivo = str(nombre_o_id).strip().lower()
    for c in cols:
        if str(c.get("name", "")).strip().lower() == objetivo:
            return c.get("id"), c.get("name")
    print("No encontre un recopilador '%s' en la encuesta %s." % (nombre_o_id, survey_id))
    print("Disponibles:")
    for c in cols:
        print("   - %s (id %s)" % (c.get("name"), c.get("id")))
    sys.exit(1)


# ----------------------------------------------------------------------------
# Utilidades de texto / valores
# ----------------------------------------------------------------------------
def limpiar_html(texto):
    if texto is None:
        return ""
    t = re.sub(r"<[^>]+>", "", str(texto))
    return t.replace("\n", " ").strip()


def to_num_si_aplica(v):
    """Convierte '10' -> 10 (int) para que las calificaciones queden numericas
    como en el export manual. Deja el resto como texto."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d+", s):
        try:
            return int(s)
        except ValueError:
            return s
    return v


def parse_fecha(iso):
    if not iso:
        return None
    try:
        # formato tipico: 2026-06-24T15:41:23+00:00 (UTC) -> hora local Monterrey
        dt = datetime.fromisoformat(str(iso).replace("Z", "+00:00")).replace(tzinfo=None)
        return dt + timedelta(hours=OFFSET_HORAS)
    except Exception:
        return iso


def norm_txt(v):
    """El export de SurveyMonkey convierte saltos de linea en espacios."""
    if isinstance(v, str):
        return v.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return v


def _norm_other(other):
    """La clave 'other' en details puede ser dict o lista. Devuelve lista de dicts."""
    if not other:
        return []
    if isinstance(other, dict):
        return [other]
    if isinstance(other, list):
        return other
    return []


# ----------------------------------------------------------------------------
# Construir el ESQUEMA de columnas (replica el export de SurveyMonkey)
# ----------------------------------------------------------------------------
# Columnas fijas de metadata (identicas al export manual)
META_COLS = [
    "respondent_id", "collector_id", "date_created", "date_modified",
    "ip_address", "email_address", "first_name", "last_name", "custom_1",
]


def construir_esquema(detalle):
    """
    Devuelve:
      columnas: lista de descriptores de columna, en el ORDEN del export.
      mapas_q: {qid: {'choices':{id:txt}, 'rows':{id:txt}, 'others':{id:txt}}}
      header1, header2: listas de texto para las 2 filas de encabezado.
    """
    columnas = []
    mapas_q = {}
    header1 = []
    header2 = []

    # metadata
    for m in META_COLS:
        columnas.append({"kind": "meta", "name": m})
        header1.append(m)
        header2.append("")

    for pagina in detalle.get("pages", []):
        for q in pagina.get("questions", []):
            qid = q.get("id")
            family = q.get("family")
            subtype = q.get("subtype")
            heading = ""
            if q.get("headings"):
                heading = limpiar_html(q["headings"][0].get("heading", ""))
            answers = q.get("answers", {}) or {}

            choices = answers.get("choices", []) or []
            rows = answers.get("rows", []) or []
            others = _norm_other(answers.get("other"))

            # mapas de id->texto para leer respuestas
            mapas_q[qid] = {
                "choices": {c.get("id"): limpiar_html(c.get("text", "")) for c in choices},
                "rows": {r.get("id"): limpiar_html(r.get("text", "")) for r in rows},
                "others": {o.get("id"): limpiar_html(o.get("text", "Otro")) for o in others},
            }

            columnas_antes = len(columnas)

            if family == "presentation":
                continue

            elif family == "single_choice":
                columnas.append({"kind": "single_resp", "qid": qid})
                header2.append("Response")
                for o in others:
                    columnas.append({"kind": "other_text", "qid": qid, "other_id": o.get("id")})
                    header2.append(limpiar_html(o.get("text", "Otro (especifique)")))

            elif family == "multiple_choice":
                for c in choices:
                    columnas.append({"kind": "check_choice", "qid": qid,
                                     "choice_id": c.get("id"),
                                     "choice_text": limpiar_html(c.get("text", ""))})
                    header2.append(limpiar_html(c.get("text", "")))
                for o in others:
                    columnas.append({"kind": "other_text", "qid": qid, "other_id": o.get("id")})
                    header2.append(limpiar_html(o.get("text", "Otro (especifique)")))

            elif family == "matrix":
                for r in rows:
                    columnas.append({"kind": "matrix_row", "qid": qid, "row_id": r.get("id")})
                    header2.append(limpiar_html(r.get("text", "")))
                for o in others:
                    columnas.append({"kind": "other_text", "qid": qid, "other_id": o.get("id")})
                    header2.append(limpiar_html(o.get("text", "Otro (especifique)")))

            elif family == "demographic":
                for r in rows:
                    columnas.append({"kind": "demo_row", "qid": qid, "row_id": r.get("id")})
                    header2.append(limpiar_html(r.get("text", "")))

            elif family == "open_ended":
                if subtype == "single" or not rows:
                    columnas.append({"kind": "open_single", "qid": qid})
                    header2.append("Open-Ended Response")
                else:
                    # multi (varias casillas de texto) -> una columna por fila
                    for r in rows:
                        columnas.append({"kind": "demo_row", "qid": qid, "row_id": r.get("id")})
                        header2.append(limpiar_html(r.get("text", "")))

            elif family == "datetime":
                columnas.append({"kind": "open_single", "qid": qid})
                header2.append("Response")

            else:
                # familia desconocida: intentar como pregunta simple
                columnas.append({"kind": "open_single", "qid": qid})
                header2.append("Response")

            # header1: el heading va en la 1a columna del bloque, resto vacio
            nuevas = len(columnas) - columnas_antes
            if nuevas > 0:
                header1.append(heading)
                header1.extend([""] * (nuevas - 1))

    return columnas, mapas_q, header1, header2


# ----------------------------------------------------------------------------
# Extraer respuestas -> filas segun el esquema
# ----------------------------------------------------------------------------
def indexar_respuesta(resp):
    """Devuelve {qid: [answers...]} y metadata de contacto."""
    por_q = {}
    for pagina in resp.get("pages", []):
        for q in pagina.get("questions", []):
            por_q[q.get("id")] = q.get("answers", []) or []
    return por_q


def valor_meta(nombre, resp):
    if nombre == "respondent_id":
        return resp.get("id")
    if nombre == "collector_id":
        return resp.get("collector_id")
    if nombre == "date_created":
        return parse_fecha(resp.get("date_created"))
    if nombre == "date_modified":
        return parse_fecha(resp.get("date_modified"))
    if nombre == "ip_address":
        return resp.get("ip_address")
    contacto = (resp.get("metadata", {}) or {}).get("contact", {}) or {}
    if nombre == "email_address":
        return (contacto.get("email", {}) or {}).get("value")
    if nombre == "first_name":
        return (contacto.get("first_name", {}) or {}).get("value")
    if nombre == "last_name":
        return (contacto.get("last_name", {}) or {}).get("value")
    if nombre == "custom_1":
        cv = resp.get("custom_variables") or {}
        if isinstance(cv, dict) and cv:
            return list(cv.values())[0]
        return None
    return None


def construir_fila(resp, columnas, mapas_q):
    por_q = indexar_respuesta(resp)
    fila = []
    for col in columnas:
        kind = col["kind"]
        if kind == "meta":
            fila.append(valor_meta(col["name"], resp))
            continue

        qid = col["qid"]
        answers = por_q.get(qid, [])
        mapa = mapas_q.get(qid, {"choices": {}, "rows": {}, "others": {}})
        val = None

        if kind == "single_resp":
            for a in answers:
                if "choice_id" in a:
                    val = mapa["choices"].get(a["choice_id"])
                    break
                # si eligio "Otro", el export pone la ETIQUETA del otro en Response
                if a.get("other_id"):
                    val = mapa["others"].get(a["other_id"])
                    break
                if "text" in a:
                    val = a["text"]
                    break
            fila.append(to_num_si_aplica(norm_txt(val)))

        elif kind == "other_text":
            oid = col["other_id"]
            for a in answers:
                if a.get("other_id") == oid and "text" in a:
                    val = a["text"]
                    break
            fila.append(norm_txt(val))

        elif kind == "check_choice":
            cid = col["choice_id"]
            for a in answers:
                if a.get("choice_id") == cid:
                    val = col["choice_text"]
                    break
            fila.append(val)

        elif kind == "matrix_row":
            rid = col["row_id"]
            for a in answers:
                if a.get("row_id") == rid and "choice_id" in a:
                    val = mapa["choices"].get(a["choice_id"])
                    break
            fila.append(to_num_si_aplica(val))

        elif kind == "demo_row":
            rid = col["row_id"]
            for a in answers:
                if a.get("row_id") == rid and "text" in a:
                    val = a["text"]
                    break
            fila.append(norm_txt(val))

        elif kind == "open_single":
            for a in answers:
                if "text" in a:
                    val = a["text"]
                    break
            fila.append(norm_txt(val))

        else:
            fila.append(None)

    return fila


# ----------------------------------------------------------------------------
# Descarga de respuestas (con paginacion)
# ----------------------------------------------------------------------------
def bajar_todas_respuestas(s, collector_id):
    url = "%s/collectors/%s/responses/bulk?per_page=100" % (API_BASE, collector_id)
    todas = []
    while url:
        data = _get(s, url)
        todas.extend(data.get("data", []))
        url = (data.get("links") or {}).get("next")
    return todas


# ----------------------------------------------------------------------------
# Escribir el Excel
# ----------------------------------------------------------------------------
def escribir_xlsx(ruta_salida, header1, header2, filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(header1)
    ws.append(header2)
    for fila in filas:
        ws.append(fila)
    wb.save(ruta_salida)


# ----------------------------------------------------------------------------
# Verificacion contra export manual
# ----------------------------------------------------------------------------
def verificar_contra_manual(ruta_generado, ruta_manual):
    print("\n[VERIFICAR] Comparando contra el export manual ...")
    print("   generado: %s" % ruta_generado)
    print("   manual  : %s" % ruta_manual)

    def leer_datos(ruta):
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        filas = {}
        ncols = 0
        for r in ws.iter_rows(min_row=3, values_only=True):
            r = list(r)
            ncols = max(ncols, len(r))
            if not r or r[0] is None:
                continue
            filas[str(r[0])] = r
        return filas, ncols

    gen, ncg = leer_datos(ruta_generado)
    man, ncm = leer_datos(ruta_manual)

    print("\n   Respuestas generado: %d  | manual: %d" % (len(gen), len(man)))
    print("   Columnas  generado: %d  | manual: %d" % (ncg, ncm))

    ids_g, ids_m = set(gen), set(man)
    solo_g = sorted(ids_g - ids_m)
    solo_m = sorted(ids_m - ids_g)
    if solo_g:
        print("   [!] IDs solo en generado:", solo_g[:10])
    if solo_m:
        print("   [!] IDs solo en manual   :", solo_m[:10])

    comunes = sorted(ids_g & ids_m)
    max_c = max(ncg, ncm)
    difs_por_col = {}
    total_difs = 0
    for rid in comunes:
        g = list(gen[rid]) + [None] * (max_c - len(gen[rid]))
        m = list(man[rid]) + [None] * (max_c - len(man[rid]))
        for j in range(max_c):
            if not _iguales(g[j], m[j]):
                difs_por_col[j] = difs_por_col.get(j, 0) + 1
                total_difs += 1

    print("\n   IDs en ambos: %d" % len(comunes))
    if total_difs == 0 and not solo_g and not solo_m and ncg == ncm:
        print("\n   RESULTADO: IDENTICO. El archivo por API coincide 100% con el export manual.")
        return True
    print("\n   Diferencias por columna (col_index: n_celdas distintas):")
    for j in sorted(difs_por_col):
        print("      col %2d: %d" % (j, difs_por_col[j]))
    # mostrar unos ejemplos de la primera columna con diferencias
    if difs_por_col:
        jcol = sorted(difs_por_col)[0]
        print("\n   Ejemplos en col %d:" % jcol)
        n = 0
        for rid in comunes:
            g = list(gen[rid]) + [None] * (max_c - len(gen[rid]))
            m = list(man[rid]) + [None] * (max_c - len(man[rid]))
            if not _iguales(g[jcol], m[jcol]):
                print("      id %s -> API=%r | manual=%r" % (rid, g[jcol], m[jcol]))
                n += 1
                if n >= 5:
                    break
    print("\n   RESULTADO: hay diferencias (ver arriba).")
    return False


def _iguales(a, b):
    """Comparacion tolerante: None=='' , numeros str vs int, fechas."""
    def norm(x):
        if x is None:
            return ""
        if isinstance(x, datetime):
            return x.replace(microsecond=0).isoformat()
        s = str(x).strip()
        if re.fullmatch(r"-?\d+", s):
            return str(int(s))
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
            return str(f)
        except ValueError:
            pass
        return s.lower()
    return norm(a) == norm(b)


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------
def cargar_dump(survey_id, collector_id):
    fd = os.path.join(CARPETA, "dump_details_%s.json" % survey_id)
    fr = os.path.join(CARPETA, "dump_responses_%s.json" % collector_id)
    if not (os.path.exists(fd) and os.path.exists(fr)):
        print("Faltan dumps: %s / %s" % (fd, fr))
        sys.exit(1)
    with open(fd, encoding="utf-8") as f:
        detalle = json.load(f)
    with open(fr, encoding="utf-8") as f:
        respuestas = json.load(f).get("data", [])
    return detalle, respuestas


def generar(s, survey_id, lista_col, salida):
    """
    Baja respuestas de un recopilador (o varios, si se piden explicitamente por
    linea de comandos) y escribe un xlsx. Normalmente se usa con UN recopilador,
    igual que un export manual.
    lista_col: lista de tuplas (collector_id, nombre).
    """
    detalle = _get(s, "%s/surveys/%s/details" % (API_BASE, survey_id))
    titulo = detalle.get("title", "encuesta")
    print("Encuesta : %s" % titulo)

    respuestas = []
    vistos = set()
    for collector_id, nombre_col in lista_col:
        r = bajar_todas_respuestas(s, collector_id)
        nuevos = 0
        for resp in r:
            rid = str(resp.get("id"))
            if rid in vistos:
                continue
            vistos.add(rid)
            respuestas.append(resp)
            nuevos += 1
        print("Recopilador: %s (id %s) -> %d respuestas" % (nombre_col, collector_id, nuevos))

    columnas, mapas_q, header1, header2 = construir_esquema(detalle)
    filas = [construir_fila(r, columnas, mapas_q) for r in respuestas]
    if not salida:
        salida = os.path.join(CARPETA, "%s.xlsx" % titulo)
    escribir_xlsx(salida, header1, header2, filas)
    print("\nArchivo listo: %s" % salida)
    print("Columnas: %d | Filas de datos: %d" % (len(columnas), len(filas)))
    return salida


def menu_interactivo():
    s = crear_sesion(cargar_token())

    print("\n" + "=" * 60)
    print("   DESCARGAR ENCUESTAS DESDE SURVEYMONKEY - CAINTRA")
    print("=" * 60)
    print("\n¿Qué tipo de encuesta quieres descargar?\n")
    for i, t in enumerate(TIPOS_ENCUESTA, 1):
        print("  %d. %s" % (i, t["nombre"]))
    print("  %d. Otra (escribir el ID de la encuesta)" % (len(TIPOS_ENCUESTA) + 1))

    try:
        op = int(input("\n> Opción: ").strip())
    except (ValueError, EOFError):
        print("Opción no válida.")
        return

    if op == len(TIPOS_ENCUESTA) + 1:
        survey_id = input("> ID de la encuesta: ").strip()
        salida = None
    elif 1 <= op <= len(TIPOS_ENCUESTA):
        tipo = TIPOS_ENCUESTA[op - 1]
        survey_id = tipo["survey_id"]
        salida = os.path.join(CARPETA, tipo["salida"]) if tipo["salida"] else None
        if not survey_id:
            print("\nEsta encuesta aún no tiene su ID guardado en el menú.")
            survey_id = input("> Escribe el ID de la encuesta (una vez) : ").strip()
    else:
        print("Opción no válida.")
        return

    if not survey_id:
        print("Sin ID de encuesta, no puedo continuar.")
        return

    # Listar eventos (recopiladores), más recientes primero
    cols = obtener_recopiladores(s, survey_id)
    def _clave(c):
        try:
            return int(c.get("id"))
        except (TypeError, ValueError):
            return 0
    cols_ordenados = sorted(cols, key=_clave, reverse=True)
    mostrar = cols_ordenados[:20]

    print("\nEventos más recientes de esta encuesta:\n")
    for i, c in enumerate(mostrar, 1):
        tipo_c = c.get("type", "")
        print("  %2d. %-24s [%s]" % (i, c.get("name"), tipo_c))
    print("\n  (o escribe directamente el código, ej. DC240726)")
    print("  Nota: se baja UN recopilador (igual que tu export manual). Para un")
    print("        refuerzo R1, córrelo aparte y elígelo por separado.")

    eleccion = input("\n> Número o código del evento: ").strip()
    if eleccion.isdigit() and 1 <= int(eleccion) <= len(mostrar):
        c = mostrar[int(eleccion) - 1]
        collector = (c.get("id"), c.get("name"))
    else:
        collector = resolver_collector_id(s, survey_id, eleccion)

    print()
    generar(s, survey_id, [collector], salida)
    print("\nAhora corre tu script sistema_encuestas_CAINTRA.py (opción 4).\n")


def main():
    args = sys.argv[1:]
    if not args:
        menu_interactivo()
        return

    survey_id = args[0].strip()
    resto = args[1:]

    # --collectors no necesita nombre
    if "--collectors" in resto:
        s = crear_sesion(cargar_token())
        listar_recopiladores(s, survey_id)
        return

    # Flags con valor
    salida = None
    ruta_manual = None
    saltar = set()
    if "--salida" in resto:
        i = resto.index("--salida")
        salida = resto[i + 1]
        saltar.update({i, i + 1})
    if "--verificar" in resto:
        i = resto.index("--verificar")
        ruta_manual = resto[i + 1]
        saltar.update({i, i + 1})
    usar_dump = "--desde-dump" in resto

    # Todos los tokens que NO son flags ni valores de flags = recopiladores
    nombres_col = []
    for idx, t in enumerate(resto):
        if idx in saltar or t.startswith("--"):
            continue
        nombres_col.append(t)

    if not nombres_col:
        print("Falta el/los recopilador(es). Ver: python descargar_encuestas.py %s --collectors" % survey_id)
        return

    # Modo dump (solo un recopilador, para desarrollo)
    if usar_dump:
        detalle, respuestas = cargar_dump(survey_id, nombres_col[0])
        titulo = detalle.get("title", "encuesta")
        columnas, mapas_q, header1, header2 = construir_esquema(detalle)
        filas = [construir_fila(r, columnas, mapas_q) for r in respuestas]
        if not salida:
            salida = os.path.join(CARPETA, "%s.xlsx" % titulo)
        escribir_xlsx(salida, header1, header2, filas)
        print("\nArchivo generado (dump): %s | filas: %d" % (salida, len(filas)))
        return

    s = crear_sesion(cargar_token())
    lista_col = []
    for n in nombres_col:
        cid, nom = resolver_collector_id(s, survey_id, n)
        lista_col.append((cid, nom))

    # En modo --verificar no sobrescribir el export manual de referencia
    if ruta_manual and not salida:
        detalle_tmp = _get(s, "%s/surveys/%s/details" % (API_BASE, survey_id))
        salida = os.path.join(CARPETA, "_API_%s.xlsx" % detalle_tmp.get("title", "encuesta"))

    salida = generar(s, survey_id, lista_col, salida)

    if ruta_manual:
        if not os.path.isabs(ruta_manual):
            ruta_manual = os.path.join(CARPETA, ruta_manual)
        verificar_contra_manual(salida, ruta_manual)


if __name__ == "__main__":
    main()
